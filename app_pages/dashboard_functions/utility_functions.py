import streamlit as st
import pandas as pd
import numpy as np
import io
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ---- UTILITY FUNCTIONS ----
 
def go_to(page_name):
    """Helper function to navigate between pages."""
    st.session_state.page = page_name
    st.rerun()

def create_excel_download_button(weights_df, targets_df, constraints_df, exclusions_df, button_text="**Download Weights** 📥", file_name="index_weights.xlsx"):
    """
    Create a download button in Streamlit that generates an Excel file with weights on the first sheet
    and options on the second sheet.
    
    Parameters:
    weights_df (pandas.DataFrame): DataFrame containing weights data
    options_df (pandas.DataFrame): DataFrame containing options data
    button_text (str): Text to display on the button
    file_name (str): Name of the file to download
    
    Returns:
    None: Displays a download button in the Streamlit app
    """
    # Create a BytesIO buffer
    buffer = io.BytesIO()
    
    # Create a workbook and add two sheets
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Weights"
    ws2 = wb.create_sheet(title="Targets")
    ws3 = wb.create_sheet(title="Constraints")
    ws4 = wb.create_sheet(title="Exclusions")
    
    # Add weights data to first sheet
    for r in dataframe_to_rows(weights_df, index=False, header=True):
        ws1.append(r)
    
    # Add options data to second sheet
    for r in dataframe_to_rows(targets_df, index=False, header=True):
        ws2.append(r)
    
    # Add options data to third sheet
    for r in dataframe_to_rows(constraints_df, index=False, header=True):
        ws3.append(r)
    
    # Add options data to fourth sheet
    for r in dataframe_to_rows(exclusions_df, index=False, header=True):
        ws4.append(r)

    # Save the workbook to the buffer
    wb.save(buffer)
    buffer.seek(0)
    
    # Create download button
    st.sidebar.download_button(
        label=button_text,
        data=buffer,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True
    )