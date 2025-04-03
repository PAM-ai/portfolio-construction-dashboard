import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import portfolio_construction as ptf
import io
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from importlib import reload

# ---- UTILITY FUNCTIONS ----

def calculate_index_performance(review_data, prices, index_weights):
    """
    Calculate Index and Benchmark performance metrics.

    Parameters:
        review_data (DataFrame): Data with benchmark weights and close prices
        prices (DataFrame): Price history data
        index_weights (DataFrame): Index weights after constraints and exclusions

    Returns:
        tuple: (combined_data, benchmark_returns, index_returns)
    """
    # Convert dates to datetime for consistency
    review_data["Review Date"] = pd.to_datetime(review_data["Review Date"])
    index_weights["Review Date"] = pd.to_datetime(index_weights["Review Date"])

    # Merge index weights with review data
    review_data_output = review_data.merge(
        index_weights[["Review Date", "Symbol", "IndexWeights"]], 
        on=["Review Date", "Symbol"], 
        how="left"
    )
    review_data_output["IndexWeights"] = review_data_output["IndexWeights"].fillna(0)

    # Merge with prices data
    prices["Review Date"] = pd.to_datetime(prices["Review Date"])
    prices = prices.merge(
        review_data_output[["Review Date", "Symbol", "IndexWeights"]], 
        on=["Review Date", "Symbol"], 
        how="left"
    )

    # Compute daily stock returns
    prices['Return'] = prices.sort_values("Date", ascending=True).groupby('Symbol')['Close'].pct_change()
    prices = prices.dropna()  # Drop NaN returns (first date)

    # Compute weighted returns
    prices['Weighted Benchmark Return'] = prices['Return'] * prices['Weight']
    prices['Weighted Index Return'] = prices['Return'] * prices['IndexWeights']

    # Aggregate to get total return per date
    benchmark_returns = prices.groupby('Date')['Weighted Benchmark Return'].sum()
    index_returns = prices.groupby('Date')['Weighted Index Return'].sum()

    # Verify weight sums (debugging)
    benchmark_weight_sums = prices.groupby("Date")["Weight"].sum().sort_values()
    index_weight_sums = prices.groupby("Date")["IndexWeights"].sum().sort_values()
    
    if not (0.99 <= benchmark_weight_sums.mean() <= 1.01) or not (0.99 <= index_weight_sums.mean() <= 1.01):
        st.warning(f"⚠️ Weight sums may not equal 100%. Check your data.")
    
    return review_data_output, benchmark_returns, index_returns

def calculate_metrics(portfolio_returns, benchmark_returns):
    """Calculate key risk metrics"""
    # Annualization factor (assuming daily returns)
    ann_factor = 252
    
    # Active returns
    active_returns = portfolio_returns - benchmark_returns
    
    # Tracking Error
    tracking_error = np.std(active_returns) * np.sqrt(ann_factor)
    
    # Information Ratio
    information_ratio = np.mean(active_returns) * ann_factor / tracking_error
    
    # Volatility
    portfolio_vol = np.std(portfolio_returns) * np.sqrt(ann_factor)
    benchmark_vol = np.std(benchmark_returns) * np.sqrt(ann_factor)
    
    # Sharpe Ratio (assuming risk-free rate of 2%)
    rf_rate = 0.02 / ann_factor
    portfolio_sharpe = (np.mean(portfolio_returns) - rf_rate) * ann_factor / portfolio_vol
    benchmark_sharpe = (np.mean(benchmark_returns) - rf_rate) * ann_factor / benchmark_vol
    
    # Max Drawdown
    portfolio_cum_returns = (1 + portfolio_returns).cumprod()
    benchmark_cum_returns = (1 + benchmark_returns).cumprod()
    
    portfolio_drawdown = 1 - portfolio_cum_returns / portfolio_cum_returns.cummax()
    benchmark_drawdown = 1 - benchmark_cum_returns / benchmark_cum_returns.cummax()
    
    portfolio_max_drawdown = portfolio_drawdown.max()
    benchmark_max_drawdown = benchmark_drawdown.max()
    
    return {
        'tracking_error': tracking_error,
        'information_ratio': information_ratio,
        'portfolio_vol': portfolio_vol,
        'benchmark_vol': benchmark_vol,
        'portfolio_sharpe': portfolio_sharpe,
        'benchmark_sharpe': benchmark_sharpe,
        'portfolio_max_drawdown': portfolio_max_drawdown,
        'benchmark_max_drawdown': benchmark_max_drawdown,
        'drawdown_series': {
            'portfolio': portfolio_drawdown,
            'benchmark': benchmark_drawdown
        }
    }

def risk_metrics_dashboard(portfolio_returns, benchmark_returns, dates):
    """Display the risk metrics dashboard"""
    metrics = calculate_metrics(portfolio_returns, benchmark_returns)
    
    # Create columns for KPI metrics
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric(
            label="Tracking Error (%)",
            value=f"{metrics['tracking_error']*100:.2f}%"
        )
        st.metric(
            label="Information Ratio",
            value=f"{metrics['information_ratio']:.2f}"
        )
    
    with col2:
        st.metric(
            label="Portfolio Volatility",
            value=f"{metrics['portfolio_vol']*100:.2f}%",
            delta=f"{(metrics['portfolio_vol'] - metrics['benchmark_vol'])*100:.2f}%",
            delta_color="inverse"
        )
        st.metric(
            label="Benchmark Volatility",
            value=f"{metrics['benchmark_vol']*100:.2f}%"
        )
    
    with col3:
        st.metric(
            label="Portfolio Sharpe",
            value=f"{metrics['portfolio_sharpe']:.2f}",
            delta=f"{metrics['portfolio_sharpe'] - metrics['benchmark_sharpe']:.2f}",
            delta_color="normal"
        )
        st.metric(
            label="Benchmark Sharpe",
            value=f"{metrics['benchmark_sharpe']:.2f}"
        )
    
    with col4:
        st.metric(
            label="Max Drawdown (Portfolio)",
            value=f"{metrics['portfolio_max_drawdown']*100:.2f}%",
            delta=f"{(metrics['benchmark_max_drawdown'] - metrics['portfolio_max_drawdown'])*100:.2f}%",
            delta_color="normal"
        )
        st.metric(
            label="Max Drawdown (Benchmark)",
            value=f"{metrics['benchmark_max_drawdown']*100:.2f}%"
        )

def plot_index_performance(benchmark_returns, index_returns):
    """
    Generate a comparative performance chart.
    
    Parameters:
        benchmark_returns (Series): Benchmark daily returns
        index_returns (Series): Index daily returns
        
    Returns:
        plotly.Figure: Performance comparison chart
    """
    benchmark_performance = (1 + benchmark_returns).cumprod()
    index_performance = (1 + index_returns).cumprod()

    # Create DataFrame for plotting
    perf_df = pd.DataFrame({
        "Date": benchmark_performance.index,
        "Benchmark": benchmark_performance.values,
        "Index": index_performance.values
    })

    # Melt DataFrame for Plotly format
    perf_df = perf_df.melt(id_vars="Date", var_name="Portfolio", value_name="Cumulative Return")

    # Plot using Plotly Express
    fig = px.line(
        perf_df, 
        x="Date", 
        y="Cumulative Return", 
        color="Portfolio", 
        title="Cumulative Performance: Index vs. Benchmark",
        labels={"Cumulative Return": "Growth of $1 Invested", "Date": "Date"},
        template="plotly_white"
    )
    
    # Enhance plot appearance
    fig.update_layout(
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        hovermode="x unified"
    )

    return fig

def get_factors_breakdown(review_data_output, review_date, factor, percentage=False):
    """
    Generate sector breakdown of a specific factor.
    
    Parameters:
        review_data_output (DataFrame): Combined review data
        review_date (datetime): Date for analysis
        factor (str): Factor to analyze
        percentage (bool): Whether to normalize as percentage
        
    Returns:
        tuple: (benchmark_factor_by_sector, index_factor_by_sector)
    """
    # Select data for the specified review date
    review_subset = review_data_output[review_data_output["Review Date"] == review_date]
    
    # Calculate weighted factor values by sector
    bmk_factor = review_subset.groupby("Sector").apply(lambda df: (df["Weight"] * df[factor]).sum())
    index_factor = review_subset.groupby("Sector").apply(lambda df: (df["IndexWeights"] * df[factor]).sum())
    
    # Convert to percentage if requested
    if percentage:
        bmk_factor = round(100 * (bmk_factor / bmk_factor.sum(axis=0)), 2)
        index_factor = round(100 * (index_factor / index_factor.sum(axis=0)), 2)

    return bmk_factor, index_factor

def plot_factors_breakdown(bmk_item, index_item, factor_name, sorting_column):
    """
    Generate a bar chart comparing factor breakdown between benchmark and index.
    
    Parameters:
        bmk_item (Series): Benchmark factor values by sector
        index_item (Series): Index factor values by sector
        factor_name (str): Name of the factor
        sorting_column (str): Column to sort by
        
    Returns:
        plotly.Figure: Comparison chart
    """
    # Convert to DataFrame for plotting
    df_plot = pd.DataFrame({
        "Sector": bmk_item.index,
        "Benchmark": bmk_item.values,
        "Index": index_item.values
    })

    # Sort values as requested
    df_plot = df_plot.sort_values(sorting_column, ascending=False)
    
    # Prepare data for grouped bar chart
    df_plot = df_plot.melt(id_vars="Sector", var_name="Portfolio", value_name="Value")

    # Create grouped bar chart
    factors_fig = px.bar(
        df_plot,
        x="Sector",
        y="Value",
        color="Portfolio",
        barmode="group",
        title=f"{factor_name} Breakdown by Sector",
        labels={"Value": f"{factor_name}", "Sector": "Sector"},
        template="plotly_white"
    )
    
    # Improve layout
    factors_fig.update_layout(
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
    )

    return factors_fig


def calculate_intensity_decomposition(portfolio_weights_history, benchmark_weights_history, 
                                     asset_intensity_history, dates, factor_name):
    """
    Decompose intensity reduction into weight effect and asset improvement effect
    
    Parameters:
    -----------
    portfolio_weights_history: DataFrame with assets as columns, dates as index
    benchmark_weights_history: DataFrame with assets as columns, dates as index
    asset_intensity_history: DataFrame with assets as columns, dates as index
    dates: DatetimeIndex or list of dates
    factor_name: String name of the factor being analyzed (e.g., 'carbon')
    
    Returns:
    --------
    DataFrame with decomposition results
    """
    results = []
    
    # We need at least two periods to calculate changes
    if len(dates) < 2:
        return pd.DataFrame()
    
    for i in range(1, len(dates)):
        current_date = dates[i]
        prev_date = dates[i-1]
        
        # Get data for the two periods
        p_weights_curr = portfolio_weights_history.loc[current_date]
        p_weights_prev = portfolio_weights_history.loc[prev_date]
        b_weights_curr = benchmark_weights_history.loc[current_date]
        
        intensity_curr = asset_intensity_history[factor_name].loc[current_date]
        intensity_prev = asset_intensity_history[factor_name].loc[prev_date]
        
        # Calculate portfolio and benchmark intensities for both periods
        p_intensity_curr = (p_weights_curr * intensity_curr).sum()
        p_intensity_prev = (p_weights_prev * intensity_prev).sum()
        b_intensity_curr = (b_weights_curr * intensity_curr).sum()
        
        # Calculate allocation effect (weight changes only)
        # What if we kept previous intensities but used current weights?
        p_intensity_weight_effect = (p_weights_curr * intensity_prev).sum()
        allocation_effect = p_intensity_weight_effect - p_intensity_prev
        
        # Calculate selection effect (asset intensity changes only)
        # Difference between current intensity and what it would be with prior intensities
        selection_effect = p_intensity_curr - p_intensity_weight_effect
        
        # Calculate total effect
        total_effect = p_intensity_curr - p_intensity_prev
        
        # Calculate improvement vs benchmark
        benchmark_gap = p_intensity_curr - b_intensity_curr
        
        results.append({
            'date': current_date,
            'portfolio_intensity': p_intensity_curr,
            'benchmark_intensity': b_intensity_curr,
            'prior_portfolio_intensity': p_intensity_prev,
            'allocation_effect': allocation_effect,
            'selection_effect': selection_effect,
            'total_effect': total_effect,
            'benchmark_gap': benchmark_gap
        })
    
    return pd.DataFrame(results).set_index('date')

def plot_intensity_decomposition(decomposition_df, factor_name):
    """Generate visualization for intensity decomposition"""
    
    # 1. Waterfall chart showing contribution to change
    fig_waterfall = go.Figure()
    
    # Use the most recent period for the waterfall
    latest_data = decomposition_df.iloc[-1]
    
    # Create waterfall chart data
    waterfall_data = [
        ('Starting Intensity', 0, latest_data['prior_portfolio_intensity'], 'start'),
        ('Allocation Effect', latest_data['allocation_effect'], 
         latest_data['prior_portfolio_intensity'] + latest_data['allocation_effect'], 'effect'),
        ('Selection Effect', latest_data['selection_effect'], 
         latest_data['prior_portfolio_intensity'] + latest_data['allocation_effect'] + latest_data['selection_effect'], 'effect'),
        ('Current Intensity', 0, latest_data['portfolio_intensity'], 'end')
    ]
    
    # Set colors based on effects (green for reductions, red for increases)
    colors = []
    for item in waterfall_data:
        if item[3] == 'start':
            colors.append('royalblue')
        elif item[3] == 'end':
            colors.append('royalblue')
        elif item[1] < 0:
            colors.append('green')  # Reduction (good)
        else:
            colors.append('red')    # Increase (bad)
        
    # Add traces
    fig_waterfall = go.Figure(go.Waterfall(
        name=f"{factor_name} Intensity Change",
        orientation="v",
        measure=[item[3] for item in waterfall_data],
        x=[item[0] for item in waterfall_data],
        textposition="outside",
        text=[f"{item[1]:.2f}" if item[3] == 'effect' else f"{item[2]:.2f}" for item in waterfall_data],
        y=[item[1] if item[3] != 'total' else 0 for item in waterfall_data],
        connector={"line": {"color": "#404040"}},
        decreasing={"marker": {"color": "#2EC4B6"}},  # Green for negative values (means improvement)
        increasing={"marker": {"color": "#FF5A5F"}},  # Red for positive values
        totals={"marker": {"color": "#3A506B"}}  # Navy blue for totals
    ))

    fig_waterfall.update_layout(
        title=f"{factor_name} Intensity Decomposition (Latest Period)",
        showlegend=False,
        font={"color": "#333333", "family": "Arial, sans-serif"},
        margin=dict(l=20, r=20, t=60, b=20)
    )
    
    # 2. Time series chart showing allocation and selection effects over time
    fig_time = go.Figure()
    
    # Cumulative effects over time
    decomposition_df['cumulative_allocation'] = decomposition_df['allocation_effect'].cumsum()
    decomposition_df['cumulative_selection'] = decomposition_df['selection_effect'].cumsum()
    decomposition_df['cumulative_total'] = decomposition_df['total_effect'].cumsum()
    
    # Add traces
    fig_time.add_trace(go.Scatter(
        x=decomposition_df.index,
        y=decomposition_df['cumulative_allocation'],
        mode='lines',
        name='Cumulative Allocation Effect',
        line=dict(width=2, color='orange')
    ))
    
    fig_time.add_trace(go.Scatter(
        x=decomposition_df.index,
        y=decomposition_df['cumulative_selection'],
        mode='lines',
        name='Cumulative Selection Effect',
        line=dict(width=2, color='blue')
    ))
    
    fig_time.add_trace(go.Scatter(
        x=decomposition_df.index,
        y=decomposition_df['cumulative_total'],
        mode='lines',
        name='Cumulative Total Effect',
        line=dict(width=3, color='green', dash='dash')
    ))
    
    fig_time.update_layout(
        title=f"Cumulative {factor_name} Intensity Effects Over Time",
        xaxis_title="Date",
        yaxis_title=f"{factor_name} Intensity Change",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
    )
    
    # 3. Stacked area chart showing breakdown of effects
    fig_stacked = go.Figure()
    
    fig_stacked.add_trace(go.Scatter(
        x=decomposition_df.index,
        y=decomposition_df['allocation_effect'],
        mode='lines',
        name='Allocation Effect',
        line=dict(width=0.5),
        stackgroup='one',
        groupnorm='',
        fillcolor='rgba(255, 165, 0, 0.5)'  # Orange with transparency
    ))
    
    fig_stacked.add_trace(go.Scatter(
        x=decomposition_df.index,
        y=decomposition_df['selection_effect'],
        mode='lines',
        name='Selection Effect',
        line=dict(width=0.5),
        stackgroup='one',
        fillcolor='rgba(65, 105, 225, 0.5)'  # Royal blue with transparency
    ))
    
    fig_stacked.update_layout(
        title=f"{factor_name} Intensity Change Contribution Over Time",
        xaxis_title="Date",
        yaxis_title=f"{factor_name} Intensity Change",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
    )
    
    # 4. Pie chart showing percentage contribution of each effect (overall)
    total_allocation = abs(decomposition_df['allocation_effect'].sum())
    total_selection = abs(decomposition_df['selection_effect'].sum())
    total = total_allocation + total_selection
    
    if total > 0:  # Avoid division by zero
        fig_pie = go.Figure(data=[go.Pie(
            labels=['Allocation Effect (Weight Changes)', 'Selection Effect (Asset Improvements)'],
            values=[total_allocation, total_selection],
            hole=.4,
            marker_colors=["rgba(255, 165, 0, 0.7)", "rgba(65, 105, 225, 0.7)"]
        )])
        
        fig_pie.update_layout(
            title=f"Contribution to {factor_name} Intensity Reduction"
        )
    else:
        fig_pie = None
    
    # Display the charts
    st.plotly_chart(fig_waterfall, use_container_width=True)
    
    # Tabs for remaining charts
    tabs = st.tabs(["Cumulative Effects", "Period Contributions", "Overall Contribution"])
    with tabs[0]:
        st.plotly_chart(fig_time, use_container_width=True)
    with tabs[1]:
        st.plotly_chart(fig_stacked, use_container_width=True)
    with tabs[2]:
        if fig_pie:
            st.plotly_chart(fig_pie, use_container_width=True)
        else:
            st.warning("Insufficient data to calculate contribution percentages")
    
    # Add explanatory text
    with st.expander("Understanding Intensity Decomposition"):
        st.markdown("""
        ### Allocation vs Selection Effects
        
        **Allocation Effect (Weight Changes)**: This represents how much of your intensity reduction comes from 
        changing the weights of assets in your portfolio. For example, reducing exposure to high-intensity 
        sectors and increasing exposure to low-intensity sectors.
        
        **Selection Effect (Asset Improvements)**: This represents how much of your intensity reduction comes from 
        improvements in the underlying assets themselves. For example, companies reducing their carbon footprint over time.
        
        ### Interpretation Guide
        
        - If the **Allocation Effect** dominates, your strategy is primarily achieving reduction through portfolio construction.
        - If the **Selection Effect** dominates, your strategy is benefiting from improvements in the underlying assets.
        - A balanced contribution suggests a robust approach that benefits from both portfolio construction and positive 
          developments in the market.
        
        The waterfall chart shows the most recent period's breakdown, while the time series charts show how these 
        effects have evolved over your investment horizon.
        """)

def intensity_decomposition_dashboard(portfolio_weights_history, benchmark_weights_history, 
                                    asset_intensity_history, dates, factor_options):
    """Main function to display the intensity decomposition dashboard"""
    
    # Factor selection dropdown
    selected_factor = st.selectbox(
        "Select Factor for Decomposition Analysis",
        options=factor_options,
        key="decomp_factor_select"
    )
    
    # Date range selector
    start_date = st.date_input("Start Date", value=dates[0], key="decomp_start_date")
    end_date = st.date_input("End Date", value=dates[-1], key="decomp_end_date")
    
    start_date = pd.Timestamp(start_date)
    end_date = pd.Timestamp(end_date)

    # Filter data by date range
    filtered_dates = [d for d in dates if start_date <= d <= end_date]
    
    if len(filtered_dates) >= 2:
        # Calculate decomposition
        decomposition_results = calculate_intensity_decomposition(
            portfolio_weights_history, 
            benchmark_weights_history,
            asset_intensity_history.loc[:, selected_factor] if isinstance(asset_intensity_history, pd.DataFrame) else asset_intensity_history,
            filtered_dates,
            selected_factor
        )
        
        print(decomposition_results.head())
        
        # Plot results
        if not decomposition_results.empty:
            plot_intensity_decomposition(decomposition_results, selected_factor)
        else:
            st.warning("Insufficient data to perform decomposition analysis")
    else:
        st.warning("Please select a date range with at least two data points")

# ---- APP LAYOUT AND FUNCTIONALITY ----

def index_generation_page():
    """Page for running the optimization and generating the index."""
    # Add page styling
    st.markdown("""
    <style>
        .block-container {max-width: 1600px}
        .stButton > button {width: 100%}
        div[data-testid="stMetricValue"] {font-size: 1.5rem}
        .review-data {height: 300px; overflow-y: scroll}
    </style>
    """, unsafe_allow_html=True)

    # Reload proprietary package for fresh data
    reload(ptf)

    # Ensure data exists before proceeding
    if "review_data" not in st.session_state or "prices" not in st.session_state:
        st.error("⚠️ Please upload and configure portfolio data in Step 1 first.")
        st.stop()

    review_data = st.session_state["review_data"]
    prices = st.session_state["prices"]

    st.header("Step 2: Generate Index")

    # Check for constraints
    constraints = st.session_state.get("constraints", {})
    if not constraints:
        st.warning("⚠️ No constraints set. Please add a constraint and choose your settings before navigating to this page.")
        st.stop()

    # Gather configuration parameters
    review_dates = sorted(review_data["Review Date"].unique())
    sustainable_factors = list(constraints.keys())
    excluded = st.session_state.get("excluded_sub_sectors", [])
    config = st.session_state.get("config", {})
    
    # Extract target levels and trajectory rates
    target_levels = np.array([v["target_reduction"] for v in constraints.values()])
    trajectory_rates = np.array([v["annual_reduction_rate"] for v in constraints.values()])
    
    with st.spinner("Generating index weights..."):
        # Get targets trajectory
        targets_df = ptf.get_targets(
            review_data, 
            sustainable_factors, 
            target_levels, 
            trajectory_rates, 
            reviews_per_year=2
        )
        
        # Get optimized weights
        index_weights, achieved_targets_df = ptf.get_weights(
            review_dates, 
            review_data, 
            sustainable_factors, 
            excluded, 
            targets_df, 
            "exp", 
            config
        )

    # Calculate performance metrics
    with st.spinner("Calculating performance..."):
        review_data_output, benchmark_returns, index_returns = calculate_index_performance(
            review_data, 
            prices, 
            index_weights
        )
        
        # Calculate weights history for the decomposition analysis
        # Store portfolio and benchmark weights history across time for each asset
        portfolio_weights_history = get_weights_history(review_data_output, 'IndexWeights')
        benchmark_weights_history = get_weights_history(review_data_output, 'Weight')
        
        # Extract intensity history for each asset and factor
        asset_intensity_history = {}
        for factor in sustainable_factors:
            asset_intensity_history[factor] = get_asset_intensity_history(review_data_output, factor)

    # Recap constraints
    st.sidebar.subheader("Settings Recap")
    constraints_df = pd.DataFrame({
        "Constraint": sustainable_factors,
        "Reduction": [round(-100 * value) for value in target_levels],
        "Trajectory": [round(-100 * value) for value in trajectory_rates]
    })

    # Set "Constraint" column as index (if needed for display)
    constraints_df.set_index('Constraint', inplace=True)

    # Display in sidebar as a table
    st.sidebar.table(constraints_df)
    
    weights_constraints_df = pd.DataFrame(list(config.items()), columns=["Constraint", "Value"])
    weights_constraints_df.set_index("Constraint", inplace=True)
    st.sidebar.table(weights_constraints_df)

    st.sidebar.subheader("SubSectors excluded")
    # Create a Markdown string for bullet points
    excluded_bullet_points = '\n'.join([f"- {item}" for item in excluded])
    if excluded_bullet_points:
        st.sidebar.markdown(excluded_bullet_points) # Display in Streamlit
    else:
        st.sidebar.write("No Subsectors excluded for this index")

    st.sidebar.subheader("Happy with your results ? Export the weights below")

    export_weights_button = create_excel_download_button(index_weights, constraints_df.reset_index(), weights_constraints_df.reset_index(), pd.DataFrame(excluded))

    # Display tabs for easier navigation
    tab1, tab2 = st.tabs(["📈 Financial", "🌱 Sustainable"])

    # Financial Analysis Tab
    with tab1:
        
        st.subheader("Performance Metrics")
        
        # INTEGRATION POINT 1: Risk Metrics Dashboard - already correctly placed
        risk_metrics_dashboard(index_returns, benchmark_returns, prices["Date"])
    
        st.subheader("Portfolio Performance")
        index_perf_fig = plot_index_performance(benchmark_returns, index_returns)
        st.plotly_chart(index_perf_fig, use_container_width=True)

       

    # Sustainable Analysis Tab
    with tab2:
        st.subheader("Sustainability Analysis")
        
        st.markdown("#### Sustainability Targets Evolution")
        
        factor_options = ["All Factors"] + sustainable_factors
        selected_item_trajectory = st.selectbox(
            "Select factor for trajectory analysis:",
            options=factor_options,
            key="trajectory_selector"
        )
        
        if selected_item_trajectory == "All Factors":
            df_melt = achieved_targets_df.melt(id_vars="Review Date", var_name="Metric", value_name="Value")
            fig_title = "Evolution of All Factor Intensities Over Time"
        else:
            df_melt = achieved_targets_df.melt(id_vars="Review Date", var_name="Metric", value_name="Value")
            df_melt = df_melt[df_melt["Metric"].str.contains(selected_item_trajectory)]
            fig_title = f"Evolution of {selected_item_trajectory} Intensity Over Time"

        fig = px.line(
            df_melt, 
            x="Review Date", 
            y="Value", 
            color="Metric", 
            title=fig_title,
            labels={"Value": "Factor Intensity", "Review Date": "Date"},
            template="plotly_white"
        )
        fig.update_layout(legend=dict(orientation="v", yanchor="bottom", y=1.02, xanchor="right", x=1))
        st.plotly_chart(fig, use_container_width=True)
        
        # INTEGRATION POINT 2: Add Time-Series Decomposition section
        st.markdown("#### Intensity Reduction Decomposition")
        
        # Convert review dates to datetime if not already
        review_dates_dt = pd.to_datetime(review_dates)
        
        # Prepare data for decomposition analysis
        intensity_decomposition_dashboard(
            portfolio_weights_history,
            benchmark_weights_history,
            asset_intensity_history,
            review_dates_dt,
            sustainable_factors
        )
        
        st.markdown("#### Sector Breakdown Analysis")
        
        col1, col2 = st.columns(2)
        with col1:
            selected_date = st.selectbox(
                "Select review date:", 
                options=sorted(review_data["Review Date"].unique()),
                key="date_selector"
            )
        
        with col2:
            selected_factor = st.selectbox(
                "Select factor:", 
                options=sustainable_factors,
                key="factor_selector"
            )
        
        col3, col4 = st.columns(2)
        with col3:
            show_as_percentage = st.checkbox("Show as percentage of total", value=False)
        
        with col4:
            sort_by_bmk = st.checkbox("Sort by benchmark values", value=True)
        
        # Calculate and display factor breakdown
        bmk_factor, index_factor = get_factors_breakdown(
            review_data_output, 
            selected_date, 
            selected_factor, 
            show_as_percentage
        )
        
        sorting_column = "Benchmark" if sort_by_bmk else "Index"
        factors_fig = plot_factors_breakdown(bmk_factor, index_factor, selected_factor, sorting_column)
        st.plotly_chart(factors_fig, use_container_width=True)

# Helper functions needed for the integration

def get_weights_history(review_data, weight_column):
    """
    Create a DataFrame with weights history per asset over time
    
    Parameters:
    review_data: DataFrame with review data including symbols and weights
    weight_column: Column name for the weights to extract (IndexWeights or Weight)
    
    Returns:
    DataFrame with dates as index, symbols as columns, weights as values
    """
    # Pivot the data to get symbols as columns and dates as rows
    weights_history = review_data.pivot_table(
        index='Review Date', 
        columns='Symbol', 
        values=weight_column,
        aggfunc='first'  # In case there are duplicates
    ).fillna(0)
    
    return weights_history

def get_asset_intensity_history(review_data, factor):
    """
    Create a DataFrame with intensity history per asset over time for a specific factor
    
    Parameters:
    review_data: DataFrame with review data including symbols and factor intensities
    factor: Name of the factor to extract intensity for
    
    Returns:
    DataFrame with dates as index, symbols as columns, intensities as values
    """
    factor_column = f"{factor}_Intensity" if f"{factor}_Intensity" in review_data.columns else factor
    
    # Pivot the data to get symbols as columns and dates as rows
    intensity_history = review_data.pivot_table(
        index='Review Date', 
        columns='Symbol', 
        values=factor_column,
        aggfunc='first'  # In case there are duplicates
    ).fillna(0)
    
    return intensity_history

def create_excel_download_button(weights_df, targets_df, constraints_df, exclusions_df, button_text="Export Weights", file_name="index_weights.xlsx"):
    """
    Create a download button in Streamlit that generates an Excel file with weights on the first sheet
    and options on the second sheet.
    
    Parameters:
    weights_df (pandas.DataFrame): DataFrame containing weights data
    options_df (pandas.DataFrame): DataFrame containing options data
    button_text (str): Text to display on the button
    file_name (str): Name of the file to download
    
    Returns:
    None: Displays a download button in the Streamlit app
    """
    # Create a BytesIO buffer
    buffer = io.BytesIO()
    
    # Create a workbook and add two sheets
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Weights"
    ws2 = wb.create_sheet(title="Targets")
    ws3 = wb.create_sheet(title="Constraints")
    ws4 = wb.create_sheet(title="Exclusions")
    
    # Add weights data to first sheet
    for r in dataframe_to_rows(weights_df, index=False, header=True):
        ws1.append(r)
    
    # Add options data to second sheet
    for r in dataframe_to_rows(targets_df, index=False, header=True):
        ws2.append(r)
    
    # Add options data to third sheet
    for r in dataframe_to_rows(constraints_df, index=False, header=True):
        ws3.append(r)
    
    # Add options data to fourth sheet
    for r in dataframe_to_rows(exclusions_df, index=False, header=True):
        ws4.append(r)

    # Save the workbook to the buffer
    wb.save(buffer)
    buffer.seek(0)
    
    # Create download button
    st.sidebar.download_button(
        label=button_text,
        data=buffer,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )